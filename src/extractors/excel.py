import io
from openpyxl import load_workbook

def extract_text_from_xlsx(b: bytes) -> str:
    bio = io.BytesIO(b)
    wb = load_workbook(bio, read_only=True, data_only=True)
    ws = wb.active
    lines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        lines.append(",".join([str(c) if c is not None else "" for c in row]))
        if i > 20: break
    return "\n".join(lines)
